import json
import os
import re
import subprocess

import threading
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk

import win32com.client

from openpyxl import load_workbook


def obter_diretorio():
    import sys
    if getattr(sys, "frozen", False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


CAMINHO_CONFIG = os.path.join(obter_diretorio(), "config_rete_caminho.json")
ARQUIVO_LAYOUT = os.path.join(obter_diretorio(), "config_rete_layout.json")


CAMPOS_DATA = {"data", "data_doc"}


def serial_para_data(valor):
    try:
        num = int(valor)
        if num > 40000 and num < 60000:
            return (datetime(1899, 12, 30) + timedelta(days=num)).strftime("%d/%m/%Y")
    except (ValueError, TypeError, OverflowError):
        pass
    return str(valor)


def normalizar(valor):
    texto = "" if valor is None else str(valor)
    return re.sub(r"[^A-Z0-9]", "", texto.upper())


class AbaRETE:
    """Painel RETE. Lê e escreve na planilha selecionada pelo usuário."""

    ALIASES = {
        "serial": ("SERIAL", "NRSERIAL", "NUMEROSERIAL", "SERIE"),
        "material": ("MATERIAL", "CODIGOMATERIAL", "MAT"),
        "lote": ("LOTE", "LOT", "BATCH"),
        "equipe": ("EQUIPE", "EQUIP", "TEAM"),
        "deposito": ("DEPOSITO", "DEP", "DEPÓSITO"),
        "data": ("DATA", "DATE"),
        "centro": ("CENTRO", "CENTER", "PLANT"),
        "statsis": ("STATSIS", "STATUS_SIS"),
        "status": ("STATUS", "STATUSR", "SITUACAO", "STATUSRETE"),
        "osme": ("OSME", "NUMERO_OSME"),
        "documento": ("DOC.", "DOC", "DOCUMENTO", "NUMDOC"),
        "data_doc": ("DATA DOC", "DATADOC", "DATA_DOC", "DOCDATE"),
        "obs": ("OBS", "OBSERVACAO", "OBSERVAÇÃO"),
    }

    def __init__(self, parent_frame, label_status_global=None, paletas=None, obter_tema_atual=None):
        self.parent_frame = parent_frame
        self.label_status_global = label_status_global
        self.paletas = paletas or {}
        self.obter_tema_atual = obter_tema_atual or (lambda: "Conecta Blue")
        self.registros = []
        self.todos_registros = []
        self._colunas = {}
        self._cabecalho_linha = None
        self._planilha_nome = None
        self._caminho_planilha = None
        self._cancelar_busca = False
        self.LARGURAS = {
            "linha": 55, "serial": 150, "material": 140, "lote": 70,
            "equipe": 70, "deposito": 70, "data": 75, "centro": 65,
            "statsis": 65, "status": 75, "osme": 120,
            "documento": 100, "data_doc": 75, "obs": 180,
            "resultado": 200
        }
        self._carregar_layout()
        self._carregar_caminho()
        self._construir_interface()
        self.aplicar_tema(self.obter_tema_atual())

    def _construir_interface(self):
        container = ttk.Frame(self.parent_frame)
        container.pack(fill="both", expand=True, padx=28, pady=20)

        ttk.Label(container, text="RETE — registro de OSME e Documento", font=("Segoe UI", 14, "bold")).pack(anchor="w")
        ttk.Label(container, text="Selecione a planilha, edite ou adicione registros e salve direto no arquivo."
                  ).pack(anchor="w", pady=(3, 12))

        self.lbl_arquivo = ttk.Label(container, text="Nenhuma planilha selecionada", wraplength=980)
        self.lbl_arquivo.pack(anchor="w", pady=(0, 10))

        botoes = ttk.Frame(container)
        botoes.pack(fill="x", pady=(0, 8))
        self.btn_selecionar = tk.Button(botoes, text="📂 Selecionar planilha...", command=self._selecionar_planilha,
                                        font=("Segoe UI", 10, "bold"), bd=0, padx=12, pady=9, cursor="hand2")
        self.btn_selecionar.pack(side="left")
        self.btn_atualizar = tk.Button(botoes, text="🔄 Carregar", command=self.carregar_planilha,
                                       font=("Segoe UI", 10, "bold"), bd=0, padx=16, pady=9, cursor="hand2")
        self.btn_atualizar.pack(side="left")
        self.btn_limpar_filtros = tk.Button(botoes, text="Limpar Filtros", command=self._limpar_filtros,
                                            font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_limpar_filtros.pack(side="left", padx=(8, 0))
        self.btn_editar = tk.Button(botoes, text="Editar", command=self._abrir_formulario_edicao,
                                    font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_editar.pack(side="left", padx=(8, 0))
        self.btn_adicionar = tk.Button(botoes, text="Adicionar", command=self._abrir_formulario_adicao,
                                       font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_adicionar.pack(side="left", padx=(4, 0))
        self._modo_alternativo = tk.BooleanVar(value=False)
        self.chk_alternativo = tk.Checkbutton(botoes, text="CL04", variable=self._modo_alternativo,
                                              font=("Segoe UI", 9), bd=0, cursor="hand2")
        self.chk_alternativo.pack(side="left", padx=(14, 2))
        self.btn_procurar_ordem = tk.Button(botoes, text="Procurar ORDEM", command=self._procurar_ordem,
                                            font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_procurar_ordem.pack(side="left")
        self.btn_colocar_rete = tk.Button(botoes, text="Colocar em RETE", command=self._colocar_em_rete,
                                          font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_colocar_rete.pack(side="left", padx=(4, 0))
        self.btn_cancelar_busca = tk.Button(botoes, text="Cancelar", command=self._cancelar_busca_ordem,
                                            font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=4, cursor="hand2")
        self.btn_cancelar_busca.pack(side="left", padx=(4, 0))
        self.btn_cancelar_busca.configure(state="disabled")
        self.lbl_resumo = ttk.Label(botoes, text="")
        self.lbl_resumo.pack(side="left", padx=14)

        self.CAMPOS = ("linha", "serial", "material", "lote", "equipe", "deposito", "data", "centro", "statsis", "status", "osme", "documento", "data_doc", "obs", "resultado")
        self.TITULOS = {
            "linha": "Linha", "serial": "Serial", "material": "Material", "lote": "Lote",
            "equipe": "Equipe", "deposito": "Depósito", "data": "Data", "centro": "Centro",
            "statsis": "StatSis", "status": "Status", "osme": "OSME",
            "documento": "Doc.", "data_doc": "Data Doc", "obs": "Obs",
            "resultado": "Resultado"
        }
        self.filtros_ativos = {}

        frame_tabela = ttk.Frame(container)
        frame_tabela.pack(fill="both", expand=True)

        self.tabela = ttk.Treeview(frame_tabela, columns=self.CAMPOS, show="headings", height=16)
        for campo in self.CAMPOS:
            self.tabela.heading(campo, text=self.TITULOS[campo], command=lambda c=campo: self._abrir_filtro_coluna(c))
            self.tabela.column(campo, width=self.LARGURAS[campo], anchor="center")
        scroll_v = ttk.Scrollbar(frame_tabela, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scroll_v.set)
        self.tabela.pack(side="left", fill="both", expand=True)
        scroll_v.pack(side="right", fill="y")
        self.tabela.bind("<ButtonRelease-1>", self._salvar_layout)
        self.tabela.bind("<ButtonRelease-3>", self._salvar_layout)
        self.tabela.bind("<Double-1>", self._editar_celula)

    def carregar_planilha(self):
        if not self._caminho_planilha or not os.path.isfile(self._caminho_planilha):
            messagebox.showinfo("RETE", "Selecione uma planilha primeiro.")
            return
        self.btn_atualizar.configure(state="disabled")
        self.lbl_resumo.configure(text="Lendo planilha...")
        self._status("RETE: lendo planilha...")
        threading.Thread(target=self._ler_planilha, daemon=True).start()

    def _ler_planilha(self):
        try:
            caminho = self._caminho_planilha
            workbook = load_workbook(caminho, read_only=True, data_only=True, keep_vba=caminho.lower().endswith(".xlsm"))
            planilha, cabecalho_linha, colunas = self._encontrar_tabela(workbook)
            if not {"material", "serial"}.issubset(colunas):
                encontrados = ", ".join(colunas) or "nenhuma coluna reconhecida"
                raise ValueError("Não encontrei as colunas obrigatórias MATERIAL e SERIAL. "
                                 f"Reconhecidas: {encontrados}.")

            TODOS_CAMPOS = ["serial", "material", "lote", "equipe", "deposito", "data", "centro", "statsis", "status", "osme", "documento", "data_doc", "obs"]
            registros = []
            for numero_linha, valores in enumerate(planilha.iter_rows(min_row=cabecalho_linha + 1, values_only=True), cabecalho_linha + 1):
                if not any(valores):
                    continue
                linha_dados = {"linha": numero_linha}
                tem_dado = False
                for campo in TODOS_CAMPOS:
                    if campo in colunas and len(valores) > colunas[campo]:
                        val = valores[colunas[campo]]
                        if campo in CAMPOS_DATA and isinstance(val, (int, float)):
                            val = serial_para_data(val)
                        linha_dados[campo] = val
                        if val is not None:
                            tem_dado = True
                    else:
                        linha_dados[campo] = None
                if not tem_dado:
                    continue
                linha_dados["resultado"] = "Aguardando regra RETE"
                registros.append(linha_dados)
            workbook.close()
            self.parent_frame.after(0, lambda: self._mostrar_registros(caminho, planilha.title, registros, colunas, cabecalho_linha))
        except Exception as erro:
            self.parent_frame.after(0, lambda e=erro: self._erro_leitura(str(e)))

    def _encontrar_tabela(self, workbook):
        for planilha in workbook.worksheets:
            for linha_numero, linha in enumerate(planilha.iter_rows(max_row=20, values_only=True), 1):
                indice = {}
                for posicao, valor in enumerate(linha):
                    nome = normalizar(valor)
                    for campo, aliases in self.ALIASES.items():
                        if nome in aliases:
                            indice[campo] = posicao
                if {"material", "serial"}.issubset(indice):
                    return planilha, linha_numero, indice
        raise ValueError("Não foi localizada uma tabela com MATERIAL e SERIAL nas primeiras 20 linhas das abas.")

    def _mostrar_registros(self, caminho, aba, registros, colunas=None, cabecalho_linha=None):
        self.registros = registros
        self.todos_registros = registros
        self.filtros_ativos = {}
        if colunas is not None:
            self._colunas = colunas
        if cabecalho_linha is not None:
            self._cabecalho_linha = cabecalho_linha
        if aba is not None:
            self._planilha_nome = aba
        self._atualizar_cabecalhos()
        self._exibir_registros(registros)
        self.lbl_arquivo.configure(text=f"Planilha: {caminho} | Aba: {aba}")
        self.lbl_resumo.configure(text=f"{len(registros)} registros carregados")
        self.btn_atualizar.configure(state="normal")
        self._status(f"RETE: {len(registros)} registros carregados.")

    def _exibir_registros(self, registros):
        self.tabela.delete(*self.tabela.get_children())
        for i, reg in enumerate(registros):
            valores = ["" if reg.get(c) is None else str(reg[c]) for c in self.CAMPOS]
            tag = "even" if i % 2 == 0 else "odd"
            self.tabela.insert("", "end", values=valores, tags=(tag,))

    def _atualizar_cabecalhos(self):
        for campo in self.CAMPOS:
            texto = self.TITULOS[campo]
            if campo in self.filtros_ativos:
                texto += " \u25bc"
            self.tabela.heading(campo, text=texto)

    def _abrir_filtro_coluna(self, campo):
        if not self.todos_registros:
            return
        tem_vazio = any(
            reg.get(campo) is None or str(reg.get(campo, "") or "").strip() == ""
            for reg in self.todos_registros
        )
        valores_unicos = sorted(set(
            str(reg.get(campo, "") or "") for reg in self.todos_registros
            if reg.get(campo) is not None and str(reg.get(campo, "") or "").strip()
        ), key=str)
        if tem_vazio:
            valores_unicos.insert(0, "(Vazios)")
        if not valores_unicos:
            return

        cores = self._cores()
        popup = tk.Toplevel(self.parent_frame)
        popup.title(f"Filtrar {self.TITULOS[campo]}")
        popup.geometry("320x420")
        popup.minsize(260, 300)
        popup.transient(self.parent_frame)
        popup.grab_set()

        estilo = ttk.Style()
        estilo.configure("Popup.TFrame", background=cores["bg"])
        estilo.configure("Popup.TLabel", background=cores["bg"], foreground=cores["texto"])
        estilo.configure("PopupSearch.TEntry", fieldbackground=cores["input"],
                         foreground=cores["texto"], insertcolor=cores["texto"],
                         bordercolor=cores["container"], lightcolor=cores["container"],
                         darkcolor=cores["container"])
        estilo.configure("PopupCB.TCheckbutton", background=cores["bg"],
                         foreground=cores["texto"], focuscolor=cores["accent"])
        estilo.map("PopupCB.TCheckbutton",
                   background=[("active", cores["container"])],
                   foreground=[("active", cores["texto"])],
                   indicatorcolor=[("selected", cores["accent"]), ("!selected", cores["texto"])])
        estilo.configure("PopupBtn.TButton", background=cores["container"],
                         foreground=cores["texto"], borderwidth=0, focuscolor="none")
        estilo.map("PopupBtn.TButton",
                   background=[("active", cores["hover"])],
                   foreground=[("active", "white")])
        estilo.configure("PopupOk.TButton", background=cores["accent"],
                         foreground="white", borderwidth=0, focuscolor="none")
        estilo.map("PopupOk.TButton",
                   background=[("active", cores["hover"])],
                   foreground=[("active", "white")])

        frame_top = ttk.Frame(popup, style="Popup.TFrame", padding=8)
        frame_top.pack(fill="x")

        ttk.Label(frame_top, text=f"Filtrar {self.TITULOS[campo]}",
                  style="Popup.TLabel", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        search_var = tk.StringVar()
        search_entry = ttk.Entry(frame_top, textvariable=search_var, style="PopupSearch.TEntry",
                                 font=("Segoe UI", 9))
        search_entry.pack(fill="x", pady=(4, 6))
        search_entry.focus_set()

        frame_botoes = ttk.Frame(frame_top, style="Popup.TFrame")
        frame_botoes.pack(fill="x")

        selecionados_atuais = self.filtros_ativos.get(campo)
        if selecionados_atuais is not None:
            valores_padrao = set(v.upper() for v in selecionados_atuais)
        else:
            valores_padrao = set(v.upper() for v in valores_unicos)

        frame_lista = ttk.Frame(popup, style="Popup.TFrame")
        frame_lista.pack(fill="both", expand=True, padx=8, pady=(0, 6))

        canvas = tk.Canvas(frame_lista, bg=cores["bg"], bd=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas, style="Popup.TFrame")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        check_vars = {}

        def construir_lista(busca=""):
            for widget in inner.winfo_children():
                widget.destroy()
            check_vars.clear()
            busca = busca.upper()
            matching = [v for v in valores_unicos if busca in v.upper()][:300]
            if not matching:
                ttk.Label(inner, text="Nenhum valor encontrado",
                          style="Popup.TLabel").pack(pady=20)
                inner.update_idletasks()
                canvas.configure(scrollregion=canvas.bbox("all"))
                return
            for val in matching:
                var = tk.BooleanVar(value=(val.upper() in valores_padrao))
                check_vars[val] = var
                cb = ttk.Checkbutton(inner, text=val, variable=var,
                                     style="PopupCB.TCheckbutton")
                cb.pack(fill="x", padx=2, pady=1)
            inner.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        def atualizar_lista(*args):
            construir_lista(search_var.get())

        search_var.trace_add("write", atualizar_lista)
        construir_lista()

        def selecionar_todos():
            for var in check_vars.values():
                var.set(True)

        def limpar_todos():
            for var in check_vars.values():
                var.set(False)

        btn_selecionar = ttk.Button(frame_botoes, text="Selecionar Todos",
                                    style="PopupBtn.TButton",
                                    command=selecionar_todos)
        btn_selecionar.pack(side="left", padx=(0, 4))
        btn_limpar_tudo = ttk.Button(frame_botoes, text="Limpar Todos",
                                     style="PopupBtn.TButton",
                                     command=limpar_todos)
        btn_limpar_tudo.pack(side="left")

        frame_ok = ttk.Frame(popup, style="Popup.TFrame", padding=8)
        frame_ok.pack(fill="x")

        def aplicar():
            selecionados = [v for v, var in check_vars.items() if var.get()]
            if len(selecionados) < len(valores_unicos):
                self.filtros_ativos[campo] = selecionados
            elif campo in self.filtros_ativos:
                del self.filtros_ativos[campo]
            self._atualizar_cabecalhos()
            self._aplicar_filtros()
            popup.destroy()

        btn_ok = ttk.Button(frame_ok, text="OK", style="PopupOk.TButton",
                            command=aplicar)
        btn_ok.pack(fill="x")

        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

    def _salvar_no_original(self, dados_alterados, serial_busca):
        caminho = self._caminho_planilha
        if not caminho or not os.path.isfile(caminho):
            messagebox.showerror("RETE", "Planilha não encontrada. Selecione a planilha novamente.")
            return False
        try:
            wb = load_workbook(caminho, keep_vba=True)
            planilha, _, colunas = self._encontrar_tabela(wb)
            cabecalho = self._cabecalho_linha
            serial_alvo = str(serial_busca).strip().upper()
            linha_encontrada = None
            for linha_num in range(cabecalho + 1, planilha.max_row + 1):
                val = planilha.cell(row=linha_num, column=colunas["serial"] + 1).value
                if val is not None and str(val).strip().upper() == serial_alvo:
                    linha_encontrada = linha_num
                    break
            if linha_encontrada is None:
                messagebox.showerror("RETE", f"Serial {serial_busca} não encontrado na planilha original.")
                wb.close()
                return False
            for campo, valor in dados_alterados.items():
                if campo in colunas:
                    col = colunas[campo] + 1
                    planilha.cell(row=linha_encontrada, column=col).value = valor
            wb.save(caminho)
            wb.close()
            return True
        except PermissionError:
            messagebox.showerror("RETE", "A planilha está aberta em outro programa.\nFeche-a e tente novamente.")
            return False
        except Exception as e:
            messagebox.showerror("RETE", f"Erro ao salvar na planilha:\n{e}")
            return False

    def _adicionar_no_original(self, dados_novo):
        caminho = self._caminho_planilha
        if not caminho or not os.path.isfile(caminho):
            messagebox.showerror("RETE", "Planilha não encontrada. Selecione a planilha novamente.")
            return False
        try:
            wb = load_workbook(caminho, keep_vba=True)
            planilha, cabecalho, colunas = self._encontrar_tabela(wb)
            ultima_linha = cabecalho + 1
            for linha_num in range(cabecalho + 1, planilha.max_row + 1):
                if any(planilha.cell(row=linha_num, column=c + 1).value is not None for c in colunas.values()):
                    ultima_linha = linha_num + 1
            for campo, valor in dados_novo.items():
                if campo in colunas:
                    planilha.cell(row=ultima_linha, column=colunas[campo] + 1).value = valor
            wb.save(caminho)
            wb.close()
            return True
        except PermissionError:
            messagebox.showerror("RETE", "A planilha está aberta em outro programa.\nFeche-a e tente novamente.")
            return False
        except Exception as e:
            messagebox.showerror("RETE", f"Erro ao adicionar na planilha:\n{e}")
            return False

    def _abrir_formulario(self, registro=None, titulo="Registro"):
        acao = "Editar" if registro else "Adicionar"
        cores = self._cores()
        popup = tk.Toplevel(self.parent_frame)
        popup.title(f"{acao} - RETE")
        popup.geometry("520x620")
        popup.minsize(450, 500)
        popup.transient(self.parent_frame)
        popup.grab_set()

        frame = ttk.Frame(popup, padding=16)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"{acao} registro RETE",
                  font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 12))

        campos_form = ["serial", "material", "lote", "equipe", "deposito", "data",
                       "centro", "statsis", "status", "osme", "documento", "data_doc", "obs"]
        entries = {}
        for campo in campos_form:
            linha = ttk.Frame(frame)
            linha.pack(fill="x", pady=2)
            ttk.Label(linha, text=self.TITULOS.get(campo, campo),
                      width=14, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(registro.get(campo, "") or "") if registro else "")
            entry = ttk.Entry(linha, textvariable=var, font=("Segoe UI", 9))
            entry.pack(side="left", fill="x", expand=True, padx=(4, 0))
            entries[campo] = var

        def salvar():
            dados = {c: v.get().strip() for c, v in entries.items()}
            if not dados.get("serial"):
                messagebox.showwarning("RETE", "O campo Serial é obrigatório.", parent=popup)
                return
            if registro:
                ok = self._salvar_no_original(dados, registro["serial"])
            else:
                ok = self._adicionar_no_original(dados)
            if ok:
                popup.destroy()
                messagebox.showinfo("RETE", f"Registro {acao.lower()}do com sucesso!")
                self.carregar_planilha()
            else:
                popup.destroy()

        def cancelar():
            popup.destroy()

        botoes = ttk.Frame(frame)
        botoes.pack(fill="x", pady=(16, 0))
        btn_salvar = tk.Button(botoes, text="Salvar", command=salvar,
                               font=("Segoe UI", 10, "bold"), bd=0, padx=20, pady=8, cursor="hand2")
        btn_salvar.pack(side="right", padx=(8, 0))
        btn_cancelar = tk.Button(botoes, text="Cancelar", command=cancelar,
                                 font=("Segoe UI", 10), bd=0, padx=16, pady=8, cursor="hand2")
        btn_cancelar.pack(side="right")

        btn_salvar.configure(bg=cores["accent"], fg="white",
                             activebackground=cores["hover"], activeforeground="white")
        btn_cancelar.configure(bg=cores["container"], fg=cores["texto"],
                               activebackground=cores["hover"], activeforeground="white")

        return popup

    def _abrir_formulario_edicao(self):
        if not self.registros:
            messagebox.showinfo("RETE", "Nenhum registro carregado.")
            return
        selecao = self.tabela.selection()
        if not selecao:
            messagebox.showinfo("RETE", "Selecione um registro na tabela para editar.")
            return
        valores = self.tabela.item(selecao[0], "values")
        if not valores or len(valores) < 2:
            return
        campos_chave = ["linha", "serial", "material", "lote", "equipe", "deposito",
                        "data", "centro", "statsis", "status", "osme", "documento",
                        "data_doc", "obs", "resultado"]
        registro = {}
        for i, c in enumerate(campos_chave):
            if i < len(valores):
                registro[c] = valores[i]
        self._abrir_formulario(registro=registro, titulo="Editar registro RETE")

    def _abrir_formulario_adicao(self):
        if not self.todos_registros:
            messagebox.showinfo("RETE", "Carregue a planilha primeiro.")
            return
        self._abrir_formulario(registro=None, titulo="Adicionar registro RETE")

    def _editar_celula(self, event):
        regiao = self.tabela.identify_region(event.x, event.y)
        if regiao != "cell":
            return
        col_id = self.tabela.identify_column(event.x)
        item = self.tabela.identify_row(event.y)
        if not col_id or not item:
            return
        idx = int(col_id.replace("#", "")) - 1
        if idx >= len(self.CAMPOS):
            return
        campo = self.CAMPOS[idx]
        if campo in ("linha", "resultado"):
            return
        valores = list(self.tabela.item(item, "values"))
        if idx >= len(valores):
            return
        cores = self._cores()
        x, y, w, h = self.tabela.bbox(item, col_id)
        entry = tk.Entry(self.tabela, font=("Segoe UI", 9), bd=1, relief="solid",
                         bg=cores["input"], fg=cores["texto"],
                         insertbackground=cores["texto"],
                         highlightcolor=cores["accent"], highlightthickness=1)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, valores[idx])
        entry.focus_set()
        entry.selection_range(0, "end")

        def confirmar(evento=None):
            child = self.tabela.get_children()
            if item not in child:
                entry.destroy()
                return
            novo_valor = entry.get().strip()
            entry.destroy()
            if novo_valor == valores[idx]:
                return
            valores[idx] = novo_valor
            self.tabela.item(item, values=valores)
            child_idx = child.index(item)
            if child_idx < len(self.registros):
                self.registros[child_idx][campo] = novo_valor
                serial = self.registros[child_idx].get("serial", "")
                if serial:
                    threading.Thread(target=self._salvar_celula,
                                     args=(self.registros[child_idx], campo, novo_valor),
                                     daemon=True).start()

        def cancelar(evento=None):
            entry.destroy()

        entry.bind("<Return>", confirmar)
        entry.bind("<Escape>", cancelar)
        entry.bind("<FocusOut>", confirmar)
        item_atual = item

    def _salvar_celula(self, registro, campo, valor):
        if not registro.get("serial"):
            return
        try:
            caminho = self._caminho_planilha
            if not caminho or not os.path.isfile(caminho):
                return
            wb = load_workbook(caminho, keep_vba=True)
            planilha, _, colunas = self._encontrar_tabela(wb)
            cabecalho = self._cabecalho_linha
            serial_alvo = str(registro["serial"]).strip().upper()
            for linha_num in range(cabecalho + 1, planilha.max_row + 1):
                val = planilha.cell(row=linha_num, column=colunas["serial"] + 1).value
                if val is not None and str(val).strip().upper() == serial_alvo:
                    if campo in colunas:
                        planilha.cell(row=linha_num, column=colunas[campo] + 1).value = valor
                    break
            wb.save(caminho)
            wb.close()
            self.parent_frame.after(0, lambda: self._status(
                f"RETE: {self.TITULOS.get(campo, campo)} salvo para serial {serial_alvo}"))
        except Exception:
            pass

    def _deposito_compativel(self, reg):
        dep = str(reg.get("deposito", "") or "").strip().upper()
        if self._modo_alternativo.get():
            return "CL04" in dep
        else:
            return "CL04" not in dep

    def _conectar_sap(self):
        try:
            sap_gui = win32com.client.GetObject("SAPGUI")
            app = sap_gui.GetScriptingEngine
            conn = app.Children(0)
            session = conn.Children(0)
            return session
        except Exception as e:
            raise RuntimeError(f"Erro ao conectar no SAP: {e}")

    def _buscar_osme_sap(self, session, serial, material):
        import time
        try:
            session.findById("wnd[0]/tbar[0]/okcd").Text = "/NIQ09"
            session.findById("wnd[0]").sendVKey(0)
            time.sleep(1)
            session.findById("wnd[0]/usr/txtSERNR-LOW").Text = str(serial)
            session.findById("wnd[0]/tbar[1]/btn[8]").press()
            time.sleep(1.5)
            grid = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell")
            row_count = grid.RowCount
            material_alvo = str(material).strip().upper() if material else ""
            linha_certa = -1
            for r in range(row_count):
                mat = grid.GetCellValue(r, "MATNR")
                if mat is not None and str(mat).strip().upper() == material_alvo:
                    linha_certa = r
                    break
            if linha_certa < 0:
                return ""
            grid.setCurrentCell(linha_certa, "SERNR")
            grid.selectedRows = str(linha_certa)
            time.sleep(0.5)
            session.findById("wnd[0]/usr/tabsTABSTRIP/tabpT\06").Select()
            time.sleep(0.5)
            session.findById("wnd[0]/usr/tabsTABSTRIP/tabpT\06/ssubSUB_DATA:SAPLITO0:0122/subSUB_0122B:SAPLITO0:1221/btn%_AUTOTEXT002").press()
            time.sleep(0.5)
            tree = session.findById("wnd[0]/usr/cntlTREE_CONTAINER/shellcont/shell")
            coll = tree.GetAllNodeKeys()
            for x in range(1, coll.Length):
                key = coll.ElementAt(x)
                num = tree.GetItemText(key, "2")
                if num is not None and num.startswith("1000") and len(num) == 12:
                    return num
        except Exception:
            pass
        return ""

    def _procurar_ordem(self):
        if not self.registros:
            messagebox.showinfo("RETE", "Carregue a planilha primeiro.")
            return
        try:
            session = self._conectar_sap()
        except RuntimeError as e:
            messagebox.showerror("RETE - SAP", str(e))
            return
        linhas = self.tabela.get_children()
        if not linhas:
            messagebox.showinfo("RETE", "Nenhum registro visível na tabela.")
            return
        self._cancelar_busca = False
        self.btn_procurar_ordem.configure(state="disabled")
        self.btn_cancelar_busca.configure(state="normal")
        self._status("RETE: buscando ordens SAP...")
        threading.Thread(target=self._executar_busca_ordem, args=(session, linhas), daemon=True).start()

    def _cancelar_busca_ordem(self):
        self._cancelar_busca = True
        self.btn_cancelar_busca.configure(state="disabled", text="Cancelando...")

    def _executar_busca_ordem(self, session, linhas):
        total = len(linhas)
        processados = 0
        atualizados = 0
        ignorados_osme = 0
        ignorados_status = 0
        erros = []
        encontrados = []
        for i, item in enumerate(linhas):
            if self._cancelar_busca:
                break
            child = self.tabela.get_children()
            if item not in child:
                continue
            idx = child.index(item)
            if idx >= len(self.registros):
                continue
            reg =         self.registros[idx]
            osme_atual = str(reg.get("osme", "") or "").strip()
            if osme_atual:
                ignorados_osme += 1
                continue
            statsis_val = str(reg.get("statsis", "") or "").strip().upper()
            status_val = str(reg.get("status", "") or "").strip().upper()
            if statsis_val != "LIDI" or status_val != "ELAB":
                ignorados_status += 1
                continue
            serial = str(reg.get("serial", "") or "").strip()
            if not serial:
                continue
            processados += 1
            material = str(reg.get("material", "") or "").strip()
            osme = self._buscar_osme_sap(session, serial, material)
            if osme:
                encontrados.append((idx, serial, osme))
                if idx < len(self.registros):
                    self.registros[idx]["osme"] = osme
                    valores = list(self.tabela.item(item, "values"))
                    campo_idx = self.CAMPOS.index("osme")
                    if campo_idx < len(valores):
                        valores[campo_idx] = osme
                        self.tabela.item(item, values=valores)
                atualizados += 1
            else:
                erros.append(serial)
            self.parent_frame.after(0, lambda c=atualizados, t=processados: self.lbl_resumo.configure(
                text=f"Ordens: {c}/{t} encontrados" + (" (cancelando...)" if self._cancelar_busca else "")))
        motivo = "cancelado" if self._cancelar_busca else "concluido"
        if encontrados:
            self.parent_frame.after(0, lambda: self._perguntar_salvar(encontrados, atualizados, ignorados_osme, ignorados_status, erros, motivo))
        else:
            partes = [f"0 ordens encontradas ({motivo})"]
            if ignorados_osme:
                partes.append(f"{ignorados_osme} ja tinham OSME")
            if ignorados_status:
                partes.append(f"{ignorados_status} status invalido")
            if erros:
                partes.append(f"{len(erros)} erro(s)")
            self.parent_frame.after(0, lambda: self._finalizar_busca(" | ".join(partes)))

    def _perguntar_salvar(self, encontrados, atualizados, ignorados_osme, ignorados_status, erros, motivo):
        msg = f"Foram encontradas {len(encontrados)} ordens:\n\n"
        for _, serial, osme in encontrados[:20]:
            msg += f"{serial} -> OSME: {osme}\n"
        if len(encontrados) > 20:
            msg += f"... e mais {len(encontrados) - 20}\n"
        msg += f"\nDeseja salvar no Excel?"
        if messagebox.askyesno("RETE - Salvar ordens", msg):
            for idx, serial, osme in encontrados:
                serial_upper = serial.upper()
                try:
                    wb = load_workbook(self._caminho_planilha, keep_vba=True)
                    ws, _, colunas = self._encontrar_tabela(wb)
                    cabecalho = self._cabecalho_linha
                    for linha_num in range(cabecalho + 1, ws.max_row + 1):
                        val = ws.cell(row=linha_num, column=colunas["serial"] + 1).value
                        if val is not None and str(val).strip().upper() == serial_upper:
                            if "osme" in colunas:
                                ws.cell(row=linha_num, column=colunas["osme"] + 1).value = osme
                            break
                    wb.save(self._caminho_planilha)
                    wb.close()
                except Exception:
                    pass
            self._finalizar_busca(f"{len(encontrados)} ordens salvas no Excel ({motivo})")
        else:
            self._finalizar_busca(f"{len(encontrados)} ordens encontradas, nao salvas ({motivo})")

    def _finalizar_busca(self, msg):
        self.btn_procurar_ordem.configure(state="normal")
        self.btn_cancelar_busca.configure(state="disabled", text="Cancelar")
        self.lbl_resumo.configure(text=msg)
        self._status(f"RETE: {msg}")

    def _colocar_em_rete(self):
        if not self.registros:
            messagebox.showinfo("RETE", "Carregue a planilha primeiro.")
            return
        if self._modo_alternativo.get():
            messagebox.showinfo("RETE", "CL04 ainda não implementado.")
            return
        diretorio = obter_diretorio()
        script = None
        for nome in os.listdir(diretorio):
            if nome.lower().endswith(".vbs") and "rete" in nome.lower():
                script = os.path.join(diretorio, nome)
                break
        if not script:
            messagebox.showinfo("RETE", "Script RETE não encontrado.\nColoque um arquivo .vbs com 'RETE' no nome na pasta do programa.")
            return
        try:
            session = self._conectar_sap()
        except RuntimeError as e:
            messagebox.showerror("RETE", str(e))
            return
        linhas = self.tabela.get_children()
        if not linhas:
            messagebox.showinfo("RETE", "Nenhum registro visível na tabela.")
            return
        self._cancelar_busca = False
        self.btn_colocar_rete.configure(state="disabled", text="Processando...")
        self.btn_cancelar_busca.configure(state="normal", text="Cancelar")
        self._status("RETE: processando...")
        threading.Thread(target=self._executar_rete, args=(script, session, linhas), daemon=True).start()

    def _executar_rete(self, script, session, linhas):
        import subprocess, tempfile, time
        try:
            with open(script, "r", encoding="utf-8") as f:
                template = f.read()
        except Exception:
            self.parent_frame.after(0, lambda: self._finalizar_rete("Erro ao ler script."))
            return
        total = len(linhas)
        processados = 0
        atualizados = 0
        ignorados_centro = 0
        ignorados_status = 0
        erros = []
        resultados = []
        for i, item in enumerate(linhas):
            if self._cancelar_busca:
                break
            child = self.tabela.get_children()
            if item not in child:
                continue
            idx = child.index(item)
            if idx >= len(self.registros):
                continue
            reg = self.registros[idx]
            centro_val = str(reg.get("centro", "") or "").strip()
            if centro_val:
                ignorados_centro += 1
                continue
            statsis_val = str(reg.get("statsis", "") or "").strip().upper()
            status_val = str(reg.get("status", "") or "").strip().upper()
            if statsis_val != "LIDI" or status_val != "ELAB":
                ignorados_status += 1
                continue
            if not self._deposito_compativel(reg):
                ignorados_status += 1
                continue
            serial = str(reg.get("serial", "") or "").strip()
            if not serial:
                continue
            processados += 1
            conteudo = template.replace("{{SERIAL}}", serial)
            osme = str(reg.get("osme", "") or "").strip()
            if "{{OSME}}" in conteudo and osme:
                conteudo = conteudo.replace("{{OSME}}", osme)
            if "{{DESTINO}}" in conteudo:
                equipe = str(reg.get("equipe", "") or "").strip()
                conteudo = conteudo.replace("{{DESTINO}}", equipe if equipe else "RETE")
            tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".vbs", delete=False, encoding="utf-8")
            tmp.write(conteudo)
            tmp.close()
            try:
                resultado = subprocess.run(
                    ["cscript.exe", "//nologo", tmp.name],
                    capture_output=True, text=True, timeout=45
                )
                saida = resultado.stdout.strip()
                if saida.startswith("ERRO:"):
                    erros.append(f"{serial}: {saida}")
                elif saida == "ORDEM VAZIA":
                    if idx < len(self.registros):
                        self.registros[idx]["obs"] = "ORDEM VAZIA"
                        valores = list(self.tabela.item(item, "values"))
                        campo_idx = self.CAMPOS.index("obs")
                        if campo_idx < len(valores):
                            valores[campo_idx] = "ORDEM VAZIA"
                            self.tabela.item(item, values=valores)
                    try:
                        wb = load_workbook(self._caminho_planilha, keep_vba=True)
                        ws, _, colunas = self._encontrar_tabela(wb)
                        cabecalho = self._cabecalho_linha
                        serial_upper = serial.upper()
                        for linha_num in range(cabecalho + 1, ws.max_row + 1):
                            val = ws.cell(row=linha_num, column=colunas["serial"] + 1).value
                            if val is not None and str(val).strip().upper() == serial_upper:
                                if "obs" in colunas:
                                    ws.cell(row=linha_num, column=colunas["obs"] + 1).value = "ORDEM VAZIA"
                                break
                        wb.save(self._caminho_planilha)
                        wb.close()
                    except Exception:
                        pass
                    atualizados += 1
                elif saida:
                    resultados.append((idx, serial, saida))
                    atualizados += 1
                elif resultado.stderr.strip():
                    erros.append(f"{serial}: {resultado.stderr.strip()}")
            except subprocess.TimeoutExpired:
                erros.append(f"{serial}: tempo excedido")
            except Exception as e:
                erros.append(f"{serial}: {e}")
            finally:
                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass
            self.parent_frame.after(0, lambda c=atualizados, t=processados: self.lbl_resumo.configure(
                text=f"RETE: {c}/{t} processados"))
        if resultados:
            self.parent_frame.after(0, lambda: self._perguntar_salvar_rete(resultados, atualizados, ignorados_centro, ignorados_status, erros))
        else:
            partes = ["0 registros processados"]
            if ignorados_centro:
                partes.append(f"{ignorados_centro} ja tinham CENTRO")
            if ignorados_status:
                partes.append(f"{ignorados_status} nao sao LIDI+ELAB")
            if erros:
                partes.append(f"{len(erros)} erro(s)")
            self.parent_frame.after(0, lambda: self._finalizar_rete(" | ".join(partes)))

    def _perguntar_salvar_rete(self, resultados, atualizados, ignorados_centro, ignorados_status, erros):
        from tkinter import messagebox
        msg = f"Foram processados {len(resultados)} registros:\n\n"
        for _, serial, saida in resultados[:20]:
            msg += f"{serial} -> {saida[:60]}\n"
        if len(resultados) > 20:
            msg += f"... e mais {len(resultados) - 20}\n"
        msg += "\nSalvar alteracoes no Excel?"
        if messagebox.askyesno("RETE - Salvar", msg):
            for idx, serial, saida in resultados:
                doc = saida.strip()
                if idx < len(self.registros):
                    if doc:
                        self.registros[idx]["documento"] = doc
                serial_upper = serial.upper()
                try:
                    wb = load_workbook(self._caminho_planilha, keep_vba=True)
                    ws, _, colunas = self._encontrar_tabela(wb)
                    cabecalho = self._cabecalho_linha
                    for linha_num in range(cabecalho + 1, ws.max_row + 1):
                        val = ws.cell(row=linha_num, column=colunas["serial"] + 1).value
                        if val is not None and str(val).strip().upper() == serial_upper:
                            if doc and "documento" in colunas:
                                ws.cell(row=linha_num, column=colunas["documento"] + 1).value = doc
                            break
                    wb.save(self._caminho_planilha)
                    wb.close()
                except Exception:
                    pass
            self._atualizar_tabela_rete()
            self._finalizar_rete(f"{len(resultados)} registros salvos no Excel")
        else:
            self._finalizar_rete(f"{len(resultados)} registros processados, nao salvos")

    def _atualizar_tabela_rete(self):
        for i, reg in enumerate(self.registros):
            valores = ["" if reg.get(c) is None else str(reg[c]) for c in self.CAMPOS]
            child = self.tabela.get_children()
            if i < len(child):
                self.tabela.item(child[i], values=valores)

    def _finalizar_rete(self, msg):
        self.btn_colocar_rete.configure(state="normal", text="Colocar em RETE")
        self.btn_cancelar_busca.configure(state="disabled", text="Cancelar")
        self.lbl_resumo.configure(text=msg)
        self._status(f"RETE: {msg}")

    def _selecionar_planilha(self):
        caminho = filedialog.askopenfilename(
            title="Selecionar planilha RETE",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")]
        )
        if caminho:
            self._caminho_planilha = caminho
            self._salvar_caminho()
            self.lbl_arquivo.configure(text=f"Planilha: {caminho}")
            self.lbl_resumo.configure(text="Planilha selecionada. Clique em Carregar.")
            self.carregar_planilha()

    def _carregar_caminho(self):
        if os.path.exists(CAMINHO_CONFIG):
            try:
                with open(CAMINHO_CONFIG, "r", encoding="utf-8") as f:
                    dados = json.load(f)
                if "caminho" in dados and os.path.isfile(dados["caminho"]):
                    self._caminho_planilha = dados["caminho"]
            except Exception:
                pass

    def _salvar_caminho(self):
        try:
            with open(CAMINHO_CONFIG, "w", encoding="utf-8") as f:
                json.dump({"caminho": self._caminho_planilha}, f)
        except Exception:
            pass

    def _carregar_layout(self):
        if os.path.exists(ARQUIVO_LAYOUT):
            try:
                with open(ARQUIVO_LAYOUT, "r", encoding="utf-8") as f:
                    layout = json.load(f)
                if "larguras" in layout:
                    for campo, larg in layout["larguras"].items():
                        if campo in self.LARGURAS:
                            self.LARGURAS[campo] = larg
            except Exception:
                pass

    def _salvar_layout(self, event=None):
        try:
            larguras = {}
            for campo in self.CAMPOS:
                larguras[campo] = self.LARGURAS[campo]
            with open(ARQUIVO_LAYOUT, "w", encoding="utf-8") as f:
                json.dump({"larguras": larguras}, f, indent=2)
        except Exception:
            pass

    def _cores(self):
        tema = self.obter_tema_atual()
        return self.paletas.get(tema, next(iter(self.paletas.values()),
                                           {"bg": "#fff", "container": "#eee", "texto": "#111",
                                            "accent": "#1976d2", "hover": "#1565c0", "input": "#fff"}))

    def _aplicar_filtros(self):
        if not self.todos_registros:
            return
        if not self.filtros_ativos:
            self._exibir_registros(self.todos_registros)
            self.lbl_resumo.configure(text=f"{len(self.todos_registros)} registros exibidos")
            return
        resultado = []
        for reg in self.todos_registros:
            match = True
            for campo, selecionados in self.filtros_ativos.items():
                if "(Vazios)" in selecionados:
                    filtro_vazio = "(Vazios)" in selecionados
                    selecionados_sem_vazio = [s for s in selecionados if s != "(Vazios)"]
                else:
                    filtro_vazio = False
                    selecionados_sem_vazio = selecionados
                val = str(reg.get(campo, "") or "").upper()
                aceitos = set(v.upper() for v in selecionados_sem_vazio)
                preenchido = bool(reg.get(campo)) and str(reg.get(campo, "") or "").strip()
                if not preenchido:
                    if not filtro_vazio:
                        match = False
                        break
                else:
                    if val not in aceitos:
                        match = False
                        break
            if match:
                resultado.append(reg)
        self._exibir_registros(resultado)
        self.lbl_resumo.configure(text=f"{len(resultado)} registros exibidos (de {len(self.todos_registros)})")

    def _limpar_filtros(self):
        self.filtros_ativos = {}
        self._atualizar_cabecalhos()
        self._aplicar_filtros()

    def _erro_leitura(self, mensagem):
        self.btn_atualizar.configure(state="normal")
        self.lbl_resumo.configure(text="Falha na leitura")
        self._status("RETE: falha ao ler planilha original.")
        messagebox.showerror("RETE — leitura da planilha", mensagem)

    def aplicar_tema(self, tema_nome):
        cores = self.paletas.get(tema_nome, next(iter(self.paletas.values()), {"bg": "#fff", "container": "#eee", "texto": "#111", "accent": "#1976d2", "hover": "#1565c0"}))
        self.btn_selecionar.configure(bg=cores["container"], fg=cores["texto"], activebackground=cores["hover"], activeforeground="white")
        self.btn_atualizar.configure(bg=cores["accent"], fg="white", activebackground=cores["hover"], activeforeground="white")
        self.btn_limpar_filtros.configure(bg=cores["container"], fg=cores["texto"], activebackground=cores["hover"], activeforeground="white")
        self.btn_editar.configure(bg=cores["container"], fg=cores["texto"], activebackground=cores["hover"], activeforeground="white")
        self.btn_adicionar.configure(bg=cores["accent"], fg="white", activebackground=cores["hover"], activeforeground="white")
        self.btn_procurar_ordem.configure(bg=cores["accent"], fg="white", activebackground=cores["hover"], activeforeground="white")
        self.btn_colocar_rete.configure(bg=cores["accent"], fg="white", activebackground=cores["hover"], activeforeground="white")
        estilo = ttk.Style()
        estilo.configure("Treeview", background=cores["container"], foreground=cores["texto"],
                         fieldbackground=cores["input"], bordercolor=cores["container"],
                         relief="flat")
        estilo.map("Treeview", background=[("selected", cores["accent"])],
                   foreground=[("selected", "white")])
        try:
            r, g, b = self._hex_to_rgb(cores["bg"])
            even = f"#{min(255, r+18):02x}{min(255, g+24):02x}{min(255, b+30):02x}"
            odd = cores["bg"]
            self.tabela.tag_configure("even", background=even)
            self.tabela.tag_configure("odd", background=odd)
        except Exception:
            pass

    def _hex_to_rgb(self, hex_cor):
        hex_cor = hex_cor.lstrip("#")
        return tuple(int(hex_cor[i:i+2], 16) for i in (0, 2, 4))

    def _status(self, texto):
        if self.label_status_global:
            self.label_status_global.configure(text=f" {texto}")


def construir_aba_rete(parent_frame, label_status_global=None, paletas=None, obter_tema_atual=None):
    return AbaRETE(parent_frame, label_status_global, paletas, obter_tema_atual)
