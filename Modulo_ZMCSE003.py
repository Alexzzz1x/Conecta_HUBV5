# -*- coding: utf-8 -*-
# Modulo ZM02 - automacao SAP | Conecta Hub 3.0
# Sempre em portugues

import os
import threading
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import Calendar

import pandas as pd
import win32com.client
import pythoncom
import traceback


# ─────────────────────────────────────────────────────────────
# Regras de equipe x base permitidas
# ─────────────────────────────────────────────────────────────
# Mapeamento equipe -> base exata conforme preenchida no SAP (ZMCSE003)
EQUIPE_BASE_MAP = {
    "RCL": ["CLAND. CONECTA LESTE"],
    "CSC": ["CONECTA LESTE - CL03"],
    "CDC": ["CONECTA SUL - CL03"],
}

INSTRUCOES = (
    "COMO USAR ESTA ABA\n"
    "=" * 42 + "\n\n"
    "1. PRE-REQUISITO:\n"
    "   O SAP GUI deve estar aberto e logado.\n\n"
    "2. DATA:\n"
    "   Informe no formato DD/MM/AAAA\n"
    "   Ex: 05/08/2025\n\n"
    "3. RECURSO (codigo da equipe na planilha, ex: CSC310, CDC410):\n"
    "   Clique em 'Carregar Recursos' apos selecionar a Equipe.\n"
    "   O recurso exato vai para o campo EQUIPE no SAP.\n\n"
    "4. BASE (preenchimento automatico no SAP):\n"
    "   RCL -> CLAND. CONECTA LESTE\n"
    "   CSC -> CONECTA LESTE - CL03\n"
    "   CDC -> CONECTA SUL - CL03\n\n"
    "5. CLIQUE EM 'INICIAR AUTOMACAO'.\n\n"
    "ATENCAO:\n"
    "  - Maximo 24 seriais por lote\n"
    "    (lotes enviados automaticamente)\n"
    "  - Nao mova o mouse nem interaja\n"
    "    com o SAP durante a execucao.\n"
    "  - A planilha sera lida diretamente da rede:\n"
    "    \\\\terra\\conecta\\arquivos\\obras\\Elpa - Almox CR Maua\\CONTROLES\n"
    "  - A planilha e aberta em somente leitura\n"
    "    (nao trava o arquivo para outros usuarios).\n"
)


# ─────────────────────────────────────────────────────────────
# Logica de negocio (SAP)
# ─────────────────────────────────────────────────────────────

def validar_selecao(equipe, base):
    """Retorna True se a combinacao equipe/base for permitida."""
    return base in EQUIPE_BASE_MAP.get(equipe, [])


def obter_sessao_sap():
    """Retorna a sessao SAP ativa via Scripting Engine."""
    pythoncom.CoInitialize()
    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        if not sap_gui:
            raise RuntimeError("SAP GUI nao esta aberto.")
        app = sap_gui.GetScriptingEngine
        connection = app.Children(0)
        session = connection.Children(0)
        return session
    except Exception as e:
        raise RuntimeError(f"Falha ao conectar ao SAP: {e}")


def processar_lote_sap(lote, equipe, base, session, callback_log=None):
    """Executa a automacao SAP para um lote de ate 24 seriais.
    Replica EXATAMENTE o script gravado do usuario.
    Retorna True em caso de sucesso, False em caso de erro.
    """
    import time

    def sap_log(msg):
        if callback_log:
            callback_log(f"   [SAP] {msg}")
        print(f"   [SAP] {msg}")

    try:
        # 1. Redimensionar janela
        try:
            session.findById("wnd[0]").resizeWorkingPane(148, 38, False)
        except Exception:
            pass

        # 2. Verificar se ja esta na tela ZMCSE003 (campo grid visivel)
        precisa_navegar = False
        try:
            session.findById("wnd[0]/usr/cntlCC_GRID_INPUT/shellcont/shell")
        except Exception:
            precisa_navegar = True

        if precisa_navegar:
            sap_log("Abrindo transacao ZMCSE003 (F00025)...")
            try:
                img_ctrl = session.findById(
                    "wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell"
                )
                img_ctrl.doubleClickNode("F00025")
                time.sleep(1)
            except Exception:
                try:
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/nZMCSE003"
                    session.findById("wnd[0]").sendVKey(0)
                    time.sleep(1)
                except Exception:
                    pass

        # 3. Inserir TODOS os seriais no grid (modifyCell apenas, sem triggerModified)
        sap_log(f"Inserindo {len(lote)} serial(is) no grid...")
        grid = session.findById("wnd[0]/usr/cntlCC_GRID_INPUT/shellcont/shell")
        for idx, serial in enumerate(lote):
            grid.modifyCell(idx, "SRNR", str(serial))

        # 4. Setar currentCellColumn e triggerModified UMA UNICA VEZ apos todos os seriais
        grid.currentCellColumn = "SRNR"
        grid.triggerModified()

        # 5. Preencher campo BASE
        sap_log(f"Preenchendo BASE: '{base}'...")
        campo_base = session.findById("wnd[0]/usr/ctxtZTBMM_CAD_EQUIPE-BASE")
        campo_base.text = base

        # 6. Preencher campo EQUIPE (Recurso)
        sap_log(f"Preenchendo EQUIPE: '{equipe}'...")
        campo_equipe = session.findById("wnd[0]/usr/ctxtZTBMM_CAD_EQUIPE-EQUIPE")
        campo_equipe.text = equipe
        campo_equipe.setFocus()
        campo_equipe.caretPosition = len(equipe)

        # 7. Confirmar campos (Enter / sendVKey 0)
        sap_log("Confirmando campos (Enter)...")
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(0.5)

        # 8. Pressionar botao GEN (Gerar)
        sap_log("Pressionando botao GEN (Gerar)...")
        grid = session.findById("wnd[0]/usr/cntlCC_GRID_INPUT/shellcont/shell")
        grid.pressToolbarButton("GEN")
        time.sleep(0.5)

        # 9. Confirmar os 2 popups pos-GEN (wnd[1]/tbar[0]/btn[0] x2)
        sap_log("Confirmando popups pos-GEN...")
        for i in range(2):
            try:
                time.sleep(0.3)
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                sap_log(f"   Popup {i+1} confirmado.")
            except Exception:
                sap_log(f"   Popup {i+1} nao encontrado (ok, pode nao existir).")
                break

        time.sleep(0.5)

        # 10. Pressionar botao SAV no grid de relatorio
        sap_log("Pressionando botao SAV (Salvar)...")
        report_grid = session.findById("wnd[0]/usr/cntlCC_GRID_REPORT/shellcont/shell")
        report_grid.pressToolbarButton("SAV")
        time.sleep(0.5)

        # 11. Confirmar popup final de salvamento (btnBUTTON_2)
        sap_log("Confirmando salvamento (btnBUTTON_2)...")
        try:
            session.findById("wnd[1]/usr/btnBUTTON_2").press()
        except Exception:
            try:
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
            except Exception:
                pass

        time.sleep(0.5)
        sap_log("Lote concluido com sucesso! Tela pronta para o proximo lote.")
        return True

    except Exception as e:
        err_tb = traceback.format_exc()
        sap_log(f"ERRO no lote SAP: {e}")
        sap_log(f"[DEBUG TRACEBACK]:\n{err_tb}")
        return False





def resolver_caminho_planilha(caminho):
    """Recebe um caminho de arquivo.
    - Se for um atalho Windows (.lnk), resolve e retorna o alvo real.
    - Caso contrario, retorna o proprio caminho.
    """
    if caminho.lower().endswith(".lnk"):
        try:
            import win32com.client as wc
            shell = wc.Dispatch("WScript.Shell")
            atalho = shell.CreateShortCut(caminho)
            return atalho.Targetpath
        except Exception:
            pass  # se falhar, tenta abrir direto
    return caminho


# Caminho padrao da planilha (atalho local .lnk que aponta para a rede)
CAMINHO_RETE_PADRAO = (
    r"C:\Users\anderson.vieira\Desktop\DEV\Projeto modular"
    r"\CONTROLE_RETE ( OFICIAL ) - Atalho.lnk"
)

def listar_recursos_da_planilha(prefixo_equipe, caminho=None):
    """Le a planilha em somente leitura e retorna os codigos de recurso
    unicos da coluna D que comecem com o prefixo informado.
    Ex: prefixo 'CSC' retorna ['CSC310', 'CSC320', ...]
    """
    import openpyxl
    plan_path = resolver_caminho_planilha(caminho or CAMINHO_RETE_PADRAO)
    if not os.path.exists(plan_path):
        return []
    try:
        wb = openpyxl.load_workbook(plan_path, read_only=True, data_only=True)
        ws = wb.active
        vistos = set()
        recursos = []
        primeira = True
        for linha in ws.iter_rows(values_only=True):
            if primeira:      # pula cabecalho
                primeira = False
                continue
            val = linha[3]    # coluna D (indice 3)
            if val is not None:
                v = str(val).strip()
                if v.upper().startswith(prefixo_equipe.upper()) and v not in vistos:
                    vistos.add(v)
                    recursos.append(v)
        wb.close()
        return sorted(recursos)
    except Exception:
        return []


def processar_zm02(data_str, equipe, base, recurso,
                   callback_log=None, callback_progresso=None,
                   caminho_planilha=None):
    """Le a planilha CONTROLE_RETE ( OFICIAL ) em modo somente leitura,
    filtra por data e equipe, divide em lotes de 24 e envia ao SAP.

    Parametros:
        data_str           -- data no formato DD/MM/AAAA
        equipe             -- nome da equipe (RCL, CSC ou CDC)
        base               -- base valida para a equipe
        callback_log       -- funcao(msg) para exibir log na UI
        callback_progresso -- funcao(atual, total) para barra de progresso
        caminho_planilha   -- caminho manual (None = usa caminho padrao da rede)
    Retorna:
        dicionario com total de seriais, numero de lotes e caminho do log
    """

    def log(msg):
        if callback_log:
            callback_log(msg)
        print(msg)

    if not validar_selecao(equipe, base):
        raise ValueError(
            f"Combinacao equipe '{equipe}' + base '{base}' nao e permitida."
        )

    # Definir caminho da planilha
    plan_path = caminho_planilha or CAMINHO_RETE_PADRAO

    # Resolver atalho Windows (.lnk) se necessario
    plan_path = resolver_caminho_planilha(plan_path)

    if not os.path.exists(plan_path):
        raise FileNotFoundError(
            f"Planilha nao encontrada:\n{plan_path}\n\n"
            "Verifique se o servidor 'terra' esta acessivel\n"
            "ou selecione o arquivo manualmente."
        )

    log(f"Abrindo planilha em modo somente leitura: {plan_path}")
    import openpyxl
    wb = openpyxl.load_workbook(plan_path, read_only=True, data_only=True)

    ABA_ALVO = "RETE"
    if ABA_ALVO in wb.sheetnames:
        ws = wb[ABA_ALVO]
        log(f"Aba selecionada: '{ABA_ALVO}'")
    else:
        ws = wb.active
        log(f"Aba '{ABA_ALVO}' nao encontrada — usando aba ativa: '{ws.title}'")
        log(f"Abas disponiveis: {wb.sheetnames}")

    linhas = list(ws.values)
    wb.close()

    import datetime as _dt

    def _normalizar_data(val):
        try:
            if val is None or pd.isna(val):
                return ""
        except (TypeError, ValueError):
            pass

        if isinstance(val, (_dt.datetime, _dt.date)):
            try:
                return val.strftime("%d/%m/%Y")
            except Exception:
                return ""

        s = str(val).strip()
        if not s or s in ("None", "NaT", "nan", "NaTType"):
            return ""
        try:
            return pd.to_datetime(s, dayfirst=True).strftime("%d/%m/%Y")
        except Exception:
            return s

    cabecalho = [str(c) if c is not None else "" for c in linhas[0]]
    dados = [list(row) for row in linhas[1:]]
    df = pd.DataFrame(dados, columns=cabecalho)

    log(f"Planilha lida em modo somente leitura — {len(df)} linhas de dados.")
    log(f"Cabecalho: {list(df.columns[:8])}")

    df["_data_norm"] = df.iloc[:, 5].apply(_normalizar_data)
    df["_recurso_norm"] = df.iloc[:, 3].apply(
        lambda v: str(v).strip() if v is not None else ""
    )

    datas_amostra    = df.loc[df["_data_norm"] != "", "_data_norm"].unique()[:5].tolist()
    recursos_amostra = df.loc[df["_recurso_norm"] != "", "_recurso_norm"].unique()[:10].tolist()
    log(f"[Diag] Datas (amostra): {datas_amostra}")
    log(f"[Diag] Recursos (amostra): {recursos_amostra}")

    filtro = (
        (df["_data_norm"] == data_str)
        & (df["_recurso_norm"] == recurso.strip())
    )
    seriais = df.loc[filtro, df.columns[0]].dropna().tolist()
    seriais = [str(s).strip() for s in seriais if str(s).strip() not in ("", "None", "nan")]

    n_data = (df["_data_norm"] == data_str).sum()
    n_rec  = (df["_recurso_norm"] == recurso.strip()).sum()
    log(f"[Diag] Linhas com data '{data_str}': {n_data} | com recurso '{recurso}': {n_rec}")

    if not seriais:
        log(f"⚠️ Nenhum serial encontrado para data '{data_str}' e recurso '{recurso}'. (Linhas com essa data: {n_data} | com esse recurso: {n_rec})")
        log(f"-> Pulando recurso {recurso} e passando para o proximo...")
        return {"total": 0, "lotes": 0, "log": None, "vazio": True}


    log(f"{len(seriais)} serial(is) encontrado(s) — Recurso: {recurso} | Data: {data_str}")

    lotes = [seriais[i: i + 24] for i in range(0, len(seriais), 24)]
    log(f"Total de lotes: {len(lotes)} (maximo 24 seriais cada)")

    log("Conectando ao SAP GUI...")
    session = obter_sessao_sap()
    log("Sessao SAP obtida com sucesso.")

    resultados = []
    for idx, lote in enumerate(lotes, start=1):
        log(f"Processando lote {idx}/{len(lotes)} — {len(lote)} serial(is)...")
        if callback_progresso:
            callback_progresso(idx - 1, len(lotes))

        sucesso = processar_lote_sap(lote, recurso, base, session, callback_log=callback_log)
        status = "OK" if sucesso else "FALHOU"
        log(f"   -> Lote {idx}: {status}")
        resultados.append({"lote": idx, "qtde": len(lote), "sucesso": sucesso})

    if callback_progresso:
        callback_progresso(len(lotes), len(lotes))

    pasta = os.path.dirname(os.path.abspath(__file__))
    agora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(pasta, f"log_zm02_{agora}.xlsx")
    pd.DataFrame(resultados).to_excel(log_path, index=False)
    log(f"Log salvo em: {log_path}")

    return {"total": len(seriais), "lotes": len(lotes), "log": log_path}


def resolver_equipe_e_base(recurso):
    """Retorna tuple (equipe, base) com base no prefixo do recurso (RCL, CSC, CDC)."""
    rec = str(recurso).strip().upper()
    if rec.startswith("RCL"):
        return "RCL", "CLAND. CONECTA LESTE"
    elif rec.startswith("CSC"):
        return "CSC", "CONECTA LESTE - CL03"
    elif rec.startswith("CDC"):
        return "CDC", "CONECTA SUL - CL03"
    for eq, bases in EQUIPE_BASE_MAP.items():
        if rec.startswith(eq):
            return eq, bases[0]
    return None, None


def processar_zm02_varredura(data_inicio_str, data_fim_str,
                             equipe_filtro=None,
                             recursos_lista=None,
                             callback_log=None, callback_progresso=None,
                             callback_recurso_concluido=None,
                             callback_lista_recursos=None,
                             callback_cancelar=None,
                             caminho_planilha=None):
    """Executa a Varredura Total no periodo [data_inicio_str ... data_fim_str].
    Agrupa TODOS os seriais do periodo por RECURSO (ex: CDC410).
    Para cada recurso: envia todos os seus seriais em lotes de ate 24 para o SAP
    antes de passar para o proximo recurso.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)
        print(msg)

    dt_inicio = datetime.datetime.strptime(data_inicio_str, "%d/%m/%Y").date()
    if data_fim_str and data_fim_str.strip():
        dt_fim = datetime.datetime.strptime(data_fim_str, "%d/%m/%Y").date()
    else:
        dt_fim = dt_inicio

    if dt_inicio > dt_fim:
        raise ValueError("Data inicial nao pode ser maior que a data final.")

    str_periodo = f"{dt_inicio.strftime('%d/%m/%Y')} ate {dt_fim.strftime('%d/%m/%Y')}"
    plan_path = resolver_caminho_planilha(caminho_planilha or CAMINHO_RETE_PADRAO)
    if not os.path.exists(plan_path):
        raise FileNotFoundError(f"Planilha nao encontrada:\n{plan_path}")

    log(f"============================================================")
    log(f" INICIANDO VARREDURA POR RECURSO NO PERÍODO: {str_periodo}")
    if equipe_filtro:
        log(f" Filtro de Equipe: {equipe_filtro}")
    log(f"============================================================")
    log(f"Lendo planilha em modo somente leitura: {plan_path}")

    import openpyxl
    wb = openpyxl.load_workbook(plan_path, read_only=True, data_only=True)
    ws = wb["RETE"] if "RETE" in wb.sheetnames else wb.active
    linhas = list(ws.values)
    wb.close()

    if not linhas or len(linhas) < 2:
        raise ValueError("A planilha esta vazia ou sem dados alem do cabecalho.")

    import datetime as _dt

    def _para_data(val):
        if val is None or pd.isna(val):
            return None
        if isinstance(val, _dt.datetime):
            return val.date()
        if isinstance(val, _dt.date):
            return val
        s = str(val).strip()
        if not s or s in ("None", "NaT", "nan"):
            return None
        try:
            return pd.to_datetime(s, dayfirst=True).date()
        except Exception:
            return None

    cabecalho = [str(c) if c is not None else "" for c in linhas[0]]
    dados = [list(row) for row in linhas[1:]]
    df = pd.DataFrame(dados, columns=cabecalho)

    df["_dt_obj"] = df.iloc[:, 5].apply(_para_data)
    df["_recurso"] = df.iloc[:, 3].apply(lambda v: str(v).strip() if v is not None else "")

    mask = (df["_dt_obj"].notna()) & (df["_dt_obj"] >= dt_inicio) & (df["_dt_obj"] <= dt_fim) & (df["_recurso"] != "")
    if equipe_filtro:
        mask = mask & (df["_recurso"].str.upper().str.startswith(equipe_filtro.upper()))

    df_filtrado = df.loc[mask]

    log(f"[DEBUG] Total de linhas lidas da planilha: {len(df)}")
    log(f"[DEBUG] Registros no período ({str_periodo}): {len(df_filtrado)}")
    if equipe_filtro:
        log(f"[DEBUG] Filtro de equipe ativo: '{equipe_filtro}'")
    if recursos_lista:
        log(f"[DEBUG] Fila de recursos enviada pela UI ({len(recursos_lista)}): {', '.join(recursos_lista[:10])}{'...' if len(recursos_lista) > 10 else ''}")

    if df_filtrado.empty:
        log(f"[DEBUG] Nenhum registro bateu com os filtros de data/equipe.")
        raise ValueError(f"Nenhum registro encontrado para a configuracao no periodo de {str_periodo}.")

    if recursos_lista:
        # Mantem a ordem da lista passada (ex: comecando a partir do recurso selecionado)
        recursos_unicos = [r for r in recursos_lista if r in df_filtrado["_recurso"].values]
        log(f"[DEBUG] Recursos da fila que possuem seriais na planilha ({len(recursos_unicos)}): {', '.join(recursos_unicos)}")
    else:
        recursos_unicos = sorted(df_filtrado["_recurso"].unique())
        log(f"[DEBUG] Todos os recursos com seriais identificados no periodo ({len(recursos_unicos)}): {', '.join(recursos_unicos)}")


    tarefas = []
    total_lotes_geral = 0
    total_seriais_geral = 0
    resumo_lista = []

    for rec in recursos_unicos:
        eq, base = resolver_equipe_e_base(rec)
        if not eq or not base:
            log(f"[AVISO] Recurso '{rec}' ignorado: prefixo nao reconhecido (deve ser RCL, CSC ou CDC).")
            continue

        seriais_rec = df_filtrado.loc[df_filtrado["_recurso"] == rec, df_filtrado.columns[0]].dropna().tolist()
        seriais_rec = [str(s).strip() for s in seriais_rec if str(s).strip() not in ("", "None", "nan")]

        if not seriais_rec:
            continue

        lotes_rec = [seriais_rec[i: i + 24] for i in range(0, len(seriais_rec), 24)]
        total_lotes_geral += len(lotes_rec)
        total_seriais_geral += len(seriais_rec)

        tarefas.append({
            "recurso": rec,
            "equipe": eq,
            "base": base,
            "seriais": seriais_rec,
            "lotes": lotes_rec
        })
        resumo_lista.append({"recurso": rec, "qtd": len(seriais_rec), "lotes": len(lotes_rec)})

    if not tarefas:
        raise ValueError("Nenhum recurso valido com seriais para processamento no periodo.")

    if callback_lista_recursos:
        callback_lista_recursos(resumo_lista)

    log(f"\nResumo da Fila: {len(tarefas)} recurso(s) | {total_seriais_geral} serial(is) | {total_lotes_geral} lote(s) no total")

    for t in tarefas:
        log(f"  • {t['recurso']:<8} (Equipe: {t['equipe']}) -> {len(t['seriais'])} seriais em {len(t['lotes'])} lote(s)")

    log("\nConectando ao SAP GUI...")
    session = obter_sessao_sap()
    log("Sessao SAP obtida com sucesso.")

    lote_global = 0
    resultados = []
    recursos_sucesso = []
    recursos_falhas = []

    for idx_rec, t in enumerate(tarefas, start=1):
        if callback_cancelar and callback_cancelar():
            log("\n⏹ VARREDURA INTERROMPIDA PELO USUÁRIO!")
            break

        rec = t["recurso"]
        eq = t["equipe"]
        base = t["base"]
        lotes = t["lotes"]
        total_rec_seriais = len(t['seriais'])

        log(f"\n============================================================")
        log(f"[{idx_rec}/{len(tarefas)}] PROCESSANDO RECURSO: {rec}")
        log(f"   Equipe: {eq} | Base: {base}")
        log(f"   Total no período ({str_periodo}): {total_rec_seriais} seriais em {len(lotes)} lote(s)")
        log(f"============================================================")

        rec_ok = True
        try:
            for idx_lote, lote in enumerate(lotes, start=1):
                if callback_cancelar and callback_cancelar():
                    log("⏹ Lote cancelado pelo usuário.")
                    rec_ok = False
                    break

                lote_global += 1
                log(f"   -> Envia Lote {idx_lote}/{len(lotes)} de {rec} ({len(lote)} seriais)...")
                if callback_progresso:
                    callback_progresso(lote_global, total_lotes_geral)

                sucesso = processar_lote_sap(lote, rec, base, session, callback_log=callback_log)
                status = "OK" if sucesso else "FALHOU"
                log(f"   -> Lote {idx_lote}/{len(lotes)}: {status}")
                resultados.append({"recurso": rec, "lote": idx_lote, "qtde": len(lote), "sucesso": sucesso})
                if not sucesso:
                    rec_ok = False
                    log(f"⚠️ Falha no lote {idx_lote} do recurso {rec}.")
                    break
        except Exception as err_rec:
            log(f"⚠️ ERRO no processamento do recurso {rec}: {err_rec}")
            rec_ok = False

        if rec_ok:
            log(f"✔ Recurso {rec} TOTALMENTE CONCLUÍDO ({total_rec_seriais} seriais enviados).")
            recursos_sucesso.append(rec)
            if callback_recurso_concluido:
                callback_recurso_concluido(rec, "OK")
        else:
            log(f"❌ Recurso {rec} PULADO devido a erro no SAP/dados. Passando para o próximo...")
            recursos_falhas.append(rec)
            if callback_recurso_concluido:
                callback_recurso_concluido(rec, "FALHOU")

    pasta = os.path.dirname(os.path.abspath(__file__))
    agora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(pasta, f"log_zm02_varredura_{agora}.xlsx")
    pd.DataFrame(resultados).to_excel(log_path, index=False)

    log(f"\n============================================================")
    log(f" VARREDURA FINALIZADA!")
    log(f" Sucesso: {len(recursos_sucesso)} recurso(s) | Falha/Pulado: {len(recursos_falhas)} recurso(s)")
    log(f" Log completo salvo em: {log_path}")
    log(f"============================================================")

    return {
        "total_seriais": total_seriais_geral,
        "total_lotes": total_lotes_geral,
        "recursos_processados": len(tarefas),
        "recursos_sucesso": recursos_sucesso,
        "recursos_falha": recursos_falhas,
        "log": log_path
    }





# ─────────────────────────────────────────────────────────────
# Interface Tkinter
# ─────────────────────────────────────────────────────────────
def construir_aba_zmcse003(parent):
    """Constroi a aba completa ZM02 – Controle ENEL."""

    frame_principal = ttk.Frame(parent)
    frame_principal.pack(fill="both", expand=True, padx=12, pady=10)

    # ── Cabecalho ──────────────────────────────────────────
    lbl_titulo = tk.Label(
        frame_principal,
        text="  Entrada ZM02 – Controle ENEL (ZMCSE003)",
        font=("Segoe UI", 13, "bold"),
        anchor="w",
    )
    lbl_titulo.pack(fill="x", pady=(0, 2))

    ttk.Separator(frame_principal, orient="horizontal").pack(fill="x", pady=(0, 8))

    # ── Corpo: painel esquerdo + painel direito ─────────────
    frame_corpo = tk.Frame(frame_principal)
    frame_corpo.pack(fill="both", expand=True)
    frame_corpo.columnconfigure(0, weight=1)
    frame_corpo.columnconfigure(1, weight=2)
    frame_corpo.rowconfigure(0, weight=1)
    frame_corpo.grid_propagate(True)

    # ── Painel esquerdo ────────────────────────────────────
    frame_esq = ttk.LabelFrame(frame_corpo, text="  Configuracao", padding=10)
    frame_esq.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

    # ── Modo de Operacao ──────────────────────────────────
    modo_var = tk.StringVar(value="VARREDURA")

    frame_modo = ttk.Frame(frame_esq)
    frame_modo.pack(fill="x", pady=(0, 6))

    rb_varredura = ttk.Radiobutton(
        frame_modo,
        text=" Varredura Total (Período X a Y)",
        value="VARREDURA",
        variable=modo_var,
    )
    rb_varredura.pack(anchor="w")

    rb_unico = ttk.Radiobutton(
        frame_modo,
        text=" Recurso Único Especificado",
        value="UNICO",
        variable=modo_var,
    )
    rb_unico.pack(anchor="w")

    ttk.Separator(frame_esq, orient="horizontal").pack(fill="x", pady=4)

    # ── Checklist de Recursos / Status ──────────────────────
    recursos_concluidos = set()
    recursos_falhas_set = set()
    todos_recursos = []

    txt_checklist = tk.Text(
        frame_esq,
        wrap="word",
        font=("Consolas", 8),
        relief="flat",
        cursor="arrow",
        height=8,
        borderwidth=0,
        padx=4,
        pady=4,
    )
    txt_checklist.pack(fill="x", expand=False, pady=(0, 6))

    def atualizar_checklist():
        txt_checklist.config(state="normal")
        txt_checklist.delete("1.0", tk.END)
        if modo_var.get() == "VARREDURA":
            txt_checklist.insert(
                "1.0",
                "🔍 MODO VARREDURA TOTAL AUTOMÁTICA\n"
                "-------------------------------------\n"
                "Status dos Recursos no Período:\n"
            )
            if todos_recursos:
                txt_checklist.insert(tk.END, "─" * 37 + "\n")
                for r in todos_recursos:
                    if r in recursos_concluidos:
                        txt_checklist.insert(tk.END, f"  [✓] {r:<8} — CONCLUIDO\n")
                    elif r in recursos_falhas_set:
                        txt_checklist.insert(tk.END, f"  [❌] {r:<8} — FALHOU/PULADO\n")
                    else:
                        txt_checklist.insert(tk.END, f"  [  ] {r:<8} — Pendente\n")
        else:
            if not todos_recursos:
                txt_checklist.insert(
                    "1.0",
                    "📋 RECURSO ÚNICO\n"
                    "-------------------------------------\n"
                    "Selecione a Equipe e clique em 'Carregar'\n"
                    "para listar os recursos."
                )
            else:
                txt_checklist.insert("1.0", "📋 RECURSOS DA EQUIPE:\n" + "─" * 37 + "\n")
                for r in todos_recursos:
                    if r in recursos_concluidos:
                        txt_checklist.insert(tk.END, f"  [✓] {r:<8} — CONCLUIDO\n")
                    elif r in recursos_falhas_set:
                        txt_checklist.insert(tk.END, f"  [❌] {r:<8} — FALHOU/PULADO\n")
                    else:
                        txt_checklist.insert(tk.END, f"  [  ] {r:<8} — Pendente\n")
        txt_checklist.config(state="disabled")


    atualizar_checklist()

    ttk.Separator(frame_esq, orient="horizontal").pack(fill="x", pady=6)

    # Campos de entrada
    frame_campos = ttk.Frame(frame_esq)
    frame_campos.pack(fill="x")
    frame_campos.columnconfigure(1, weight=1)

    # ── Helper: Seletor de data com popup de calendario ────
    def criar_campo_data(parent, row, label_text):
        """Cria um campo de data com Entry + botao calendario."""
        ttk.Label(parent, text=label_text, anchor="w").grid(
            row=row, column=0, sticky="w", pady=4
        )
        frame_dt = ttk.Frame(parent)
        frame_dt.grid(row=row, column=1, sticky="ew", padx=(8, 0), pady=4)
        frame_dt.columnconfigure(0, weight=1)

        var_data = tk.StringVar(value=datetime.date.today().strftime("%d/%m/%Y"))
        entry = ttk.Entry(frame_dt, textvariable=var_data, font=("Segoe UI", 10),
                          state="readonly", width=12)
        entry.grid(row=0, column=0, sticky="ew")

        def abrir_calendario():
            # Parsear data atual do campo
            try:
                dt_atual = datetime.datetime.strptime(var_data.get(), "%d/%m/%Y").date()
            except Exception:
                dt_atual = datetime.date.today()

            # Criar janela popup
            popup = tk.Toplevel(parent)
            popup.title("Selecionar Data")
            popup.resizable(False, False)
            popup.grab_set()

            # Centralizar perto do botao
            popup.geometry(f"+{entry.winfo_rootx()}+{entry.winfo_rooty() + 30}")

            cal = Calendar(
                popup,
                selectmode="day",
                year=dt_atual.year,
                month=dt_atual.month,
                day=dt_atual.day,
                locale="pt_BR",
                date_pattern="dd/mm/yyyy",
                font=("Segoe UI", 10),
                background="#f57c00",
                foreground="white",
                headersbackground="#e65100",
                headersforeground="white",
                selectbackground="#f57c00",
                selectforeground="white",
                normalbackground="white",
                normalforeground="black",
                weekendbackground="#fff3e0",
                weekendforeground="#e65100",
                borderwidth=0,
                cursor="hand2",
            )
            cal.pack(padx=10, pady=10)

            def confirmar():
                var_data.set(cal.get_date())
                popup.destroy()

            tk.Button(
                popup,
                text="✔  Confirmar",
                font=("Segoe UI", 10, "bold"),
                bg="#f57c00", fg="white",
                activebackground="#e65100", activeforeground="white",
                relief="flat", cursor="hand2",
                pady=6, bd=0,
                command=confirmar,
            ).pack(fill="x", padx=10, pady=(0, 10))

        btn_cal = tk.Button(
            frame_dt, text="📅", font=("Segoe UI", 10),
            bg="#f57c00", fg="white", bd=0, width=3,
            activebackground="#e65100", relief="flat",
            cursor="hand2", command=abrir_calendario,
        )
        btn_cal.grid(row=0, column=1, padx=(4, 0))

        return var_data

    var_data_ini = criar_campo_data(frame_campos, 0, "Data Inicial:")
    var_data_fim = criar_campo_data(frame_campos, 1, "Data Final:")

    ttk.Label(frame_campos, text="Equipe:", anchor="w").grid(
        row=2, column=0, sticky="w", pady=4
    )

    combo_equipe = ttk.Combobox(
        frame_campos,
        values=list(EQUIPE_BASE_MAP.keys()),
        state="readonly",
        font=("Segoe UI", 10),
    )
    combo_equipe.grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=4)

    ttk.Label(frame_campos, text="Base:", anchor="w").grid(
        row=3, column=0, sticky="w", pady=4
    )
    combo_base = ttk.Combobox(frame_campos, state="readonly", font=("Segoe UI", 10))
    combo_base.grid(row=3, column=1, sticky="ew", padx=(8, 0), pady=4)

    ttk.Label(frame_campos, text="Recurso:", anchor="w").grid(
        row=4, column=0, sticky="w", pady=4
    )
    frame_recurso = ttk.Frame(frame_campos)
    frame_recurso.grid(row=4, column=1, sticky="ew", padx=(8, 0), pady=4)
    frame_recurso.columnconfigure(0, weight=1)

    combo_recurso = ttk.Combobox(frame_recurso, state="readonly", font=("Segoe UI", 10))
    combo_recurso.grid(row=0, column=0, sticky="ew", padx=(0, 4))

    btn_carregar_recurso = ttk.Button(frame_recurso, text="Carregar", width=8)
    btn_carregar_recurso.grid(row=0, column=1, sticky="e")

    def carregar_recursos():
        nonlocal todos_recursos
        eq = combo_equipe.get().strip()
        if not eq:
            messagebox.showwarning("Selecione a equipe", "Selecione a equipe antes de carregar os recursos.")
            return
        try:
            caminho = caminho_planilha_var.get().strip()
        except NameError:
            caminho = None
        recursos = listar_recursos_da_planilha(eq, caminho or None)
        if recursos:
            todos_recursos = recursos
            combo_recurso["values"] = recursos
            combo_recurso.set(recursos[0])
            atualizar_checklist()
        else:
            todos_recursos = []
            combo_recurso["values"] = []
            combo_recurso.set("")
            atualizar_checklist()
            messagebox.showwarning(
                "Sem recursos",
                f"Nenhum recurso encontrado para equipe '{eq}'.\n"
                "Verifique se a planilha esta acessivel."
            )

    btn_carregar_recurso.config(command=carregar_recursos)

    def atualizar_bases(event=None):
        nonlocal todos_recursos
        bases = EQUIPE_BASE_MAP.get(combo_equipe.get(), [])
        combo_base["values"] = bases
        combo_base.set(bases[0] if bases else "")
        combo_recurso["values"] = []
        combo_recurso.set("")
        todos_recursos = []
        atualizar_checklist()

    combo_equipe.bind("<<ComboboxSelected>>", atualizar_bases)

    def atualizar_estado_modo(*args):
        is_varredura = modo_var.get() == "VARREDURA"
        if is_varredura:
            combo_equipe.config(state="disabled")
            combo_base.config(state="disabled")
            combo_recurso.config(state="disabled")
            btn_carregar_recurso.config(state="disabled")
        else:
            combo_equipe.config(state="readonly")
            combo_base.config(state="readonly")
            combo_recurso.config(state="readonly")
            btn_carregar_recurso.config(state="normal")
        atualizar_checklist()


    rb_varredura.config(command=atualizar_estado_modo)
    rb_unico.config(command=atualizar_estado_modo)
    atualizar_estado_modo()

    # ── Campo: caminho da planilha ────────────────────────
    ttk.Separator(frame_esq, orient="horizontal").pack(fill="x", pady=6)

    ttk.Label(frame_esq, text="Planilha (CONTROLE_RETE):", anchor="w").pack(
        fill="x", pady=(4, 2)
    )

    frame_plan = ttk.Frame(frame_esq)
    frame_plan.pack(fill="x", pady=(0, 4))
    frame_plan.columnconfigure(0, weight=1)

    caminho_planilha_var = tk.StringVar(value=CAMINHO_RETE_PADRAO)
    entry_planilha = ttk.Entry(
        frame_plan, textvariable=caminho_planilha_var, font=("Segoe UI", 7)
    )
    entry_planilha.grid(row=0, column=0, sticky="ew", padx=(0, 4))

    def selecionar_planilha():
        caminho = filedialog.askopenfilename(
            title="Selecionar planilha CONTROLE_RETE",
            filetypes=[
                ("Excel / Atalho", "*.xlsx *.xlsm *.xls *.lnk"),
                ("Todos os arquivos", "*.*"),
            ],
            initialdir=r"C:\Users\anderson.vieira\Desktop\DEV\Projeto modular",
        )
        if caminho:
            caminho_planilha_var.set(caminho)

    ttk.Button(frame_plan, text="...", width=3, command=selecionar_planilha).grid(
        row=0, column=1, sticky="e"
    )

    # Barra de progresso
    ttk.Label(frame_esq, text="Progresso dos lotes:").pack(anchor="w", pady=(10, 2))
    progresso_var = tk.DoubleVar(value=0)
    barra_progresso = ttk.Progressbar(
        frame_esq, variable=progresso_var, maximum=100
    )
    barra_progresso.pack(fill="x")

    lbl_progresso = tk.Label(
        frame_esq, text="Aguardando...", font=("Segoe UI", 8), anchor="w"
    )
    lbl_progresso.pack(fill="x", pady=(2, 10))

    # Botoes de acao (Iniciar / Cancelar)
    frame_botoes = ttk.Frame(frame_esq)
    frame_botoes.pack(fill="x", pady=(6, 0))
    frame_botoes.columnconfigure(0, weight=3)
    frame_botoes.columnconfigure(1, weight=2)

    btn_iniciar = tk.Button(
        frame_botoes,
        text="▶  INICIAR AUTOMAÇÃO",
        font=("Segoe UI", 10, "bold"),
        bg="#f57c00",
        fg="white",
        activebackground="#e65100",
        activeforeground="white",
        relief="flat",
        cursor="hand2",
        pady=10,
        bd=0,
    )
    btn_iniciar.grid(row=0, column=0, sticky="ew", padx=(0, 4))

    btn_cancelar = tk.Button(
        frame_botoes,
        text="⏹ CANCELAR",
        font=("Segoe UI", 10, "bold"),
        bg="#d32f2f",
        fg="white",
        activebackground="#b71c1c",
        activeforeground="white",
        relief="flat",
        cursor="hand2",
        pady=10,
        bd=0,
        state="disabled",
    )
    btn_cancelar.grid(row=0, column=1, sticky="ew")

    cancelar_flag = False

    def on_cancelar():
        nonlocal cancelar_flag
        cancelar_flag = True
        adicionar_log("\n⏹ CANCELAMENTO SOLICITADO PELO USUÁRIO! Parando automação...")
        lbl_progresso.config(text="Cancelando...")

    btn_cancelar.config(command=on_cancelar)


    # ── Painel direito: log ────────────────────────────────
    frame_dir = ttk.LabelFrame(frame_corpo, text="  Log de Execucao", padding=10)
    frame_dir.grid(row=0, column=1, sticky="nsew")
    frame_dir.rowconfigure(0, weight=1)
    frame_dir.columnconfigure(0, weight=1)

    txt_log = tk.Text(
        frame_dir,
        wrap="word",
        font=("Consolas", 9),
        relief="flat",
        state="normal",
        borderwidth=0,
        padx=4,
    )
    txt_log.grid(row=0, column=0, sticky="nsew")

    scroll_log = ttk.Scrollbar(frame_dir, orient="vertical", command=txt_log.yview)
    scroll_log.grid(row=0, column=1, sticky="ns")
    txt_log.configure(yscrollcommand=scroll_log.set)

    btn_limpar = ttk.Button(
        frame_dir,
        text="Limpar log",
        command=lambda: txt_log.delete("1.0", tk.END),
    )
    btn_limpar.grid(row=1, column=0, columnspan=2, sticky="e", pady=(6, 0))

    # ── Callbacks de UI (thread-safe) ─────────────────────

    def adicionar_log(msg):
        def _run():
            txt_log.config(state="normal")
            txt_log.insert(tk.END, msg + "\n")
            txt_log.see(tk.END)
        txt_log.after(0, _run)

    def atualizar_progresso(atual, total):
        def _run():
            pct = (atual / total * 100) if total else 0
            progresso_var.set(pct)
            lbl_progresso.config(text=f"Lote {atual}/{total}  ({pct:.0f}%)")
        barra_progresso.after(0, _run)

    def recurso_status_callback(rec, status):
        def _run():
            if status == "OK":
                recursos_concluidos.add(rec)
            else:
                recursos_falhas_set.add(rec)
            atualizar_checklist()
        txt_log.after(0, _run)

    def lista_recursos_callback(resumo_lista):
        def _run():
            nonlocal todos_recursos
            todos_recursos = [item["recurso"] for item in resumo_lista]
            atualizar_checklist()
        txt_log.after(0, _run)


    def set_btn(habilitar):
        def _run():
            if habilitar:
                btn_iniciar.config(state="normal", text="▶  INICIAR AUTOMAÇÃO")
                btn_cancelar.config(state="disabled")
            else:
                btn_iniciar.config(state="disabled", text="⏳  Processando...")
                btn_cancelar.config(state="normal")
        btn_iniciar.after(0, _run)


    # ── Logica do botao ────────────────────────────────────

    def on_iniciar():
        nonlocal cancelar_flag
        cancelar_flag = False

        # Ler datas do calendario
        data_ini = var_data_ini.get().strip()
        data_fim = var_data_fim.get().strip()
        modo = modo_var.get()

        # Limpar log e iniciar
        txt_log.delete("1.0", tk.END)
        progresso_var.set(0)
        lbl_progresso.config(text="Iniciando...")
        set_btn(False)

        def _executar():
            try:
                caminho_usr = caminho_planilha_var.get().strip() or None
                if modo == "VARREDURA":
                    resultado = processar_zm02_varredura(
                        data_inicio_str=data_ini,
                        data_fim_str=data_fim,
                        callback_log=adicionar_log,
                        callback_progresso=atualizar_progresso,
                        callback_recurso_concluido=recurso_status_callback,
                        callback_lista_recursos=lista_recursos_callback,
                        callback_cancelar=lambda: cancelar_flag,
                        caminho_planilha=caminho_usr,
                    )

                    def _ok():
                        sucesso_lst = resultado.get("recursos_sucesso", [])
                        falha_lst   = resultado.get("recursos_falha", [])

                        txt_suc = "\n".join([f"  • {r}" for r in sucesso_lst]) if sucesso_lst else "  (Nenhum)"
                        txt_fal = "\n".join([f"  • {r}" for r in falha_lst]) if falha_lst else "  (Nenhum)"

                        adicionar_log(
                            f"\n============================================================\n"
                            f"📊 RELATÓRIO GERAL DA VARREDURA TOTAL\n"
                            f"============================================================\n\n"
                            f"Recursos Concluídos com Sucesso ({len(sucesso_lst)}):\n"
                            f"{txt_suc}\n\n"
                            f"Recursos com Falha / Pulados ({len(falha_lst)}):\n"
                            f"{txt_fal}\n\n"
                            f"------------------------------------------------------------\n"
                            f"Total de Seriais Processados: {resultado['total_seriais']}\n"
                            f"Total de Lotes Executados:    {resultado['total_lotes']}\n\n"
                            f"Log detalhado salvo em:\n{resultado['log']}\n"
                            f"============================================================\n"
                        )
                    btn_iniciar.after(0, _ok)

                else:
                    equipe = combo_equipe.get().strip()
                    base   = combo_base.get().strip()
                    recurso_sel = combo_recurso.get().strip()

                    if not equipe:
                        raise ValueError("Selecione a Equipe para continuar.")

                    # Se o usuario tiver a lista de recursos carregada, inicia a partir do selecionado
                    recursos_fila = None
                    if todos_recursos:
                        if recurso_sel in todos_recursos:
                            idx_start = todos_recursos.index(recurso_sel)
                            recursos_fila = todos_recursos[idx_start:]
                        else:
                            recursos_fila = todos_recursos
                    elif recurso_sel:
                        recursos_fila = [recurso_sel]

                    adicionar_log(f"\n============================================================")
                    adicionar_log(f" EXECUÇÃO EM LOOP POR RECURSO — Data: {data_ini} até {data_fim or data_ini}")
                    adicionar_log(f" Equipe: {equipe} | Recurso Inicial: {recurso_sel or 'Primeiro da fila'}")
                    adicionar_log(f"============================================================")

                    resultado = processar_zm02_varredura(
                        data_inicio_str=data_ini,
                        data_fim_str=data_fim or data_ini,
                        equipe_filtro=equipe,
                        recursos_lista=recursos_fila,
                        callback_log=adicionar_log,
                        callback_progresso=atualizar_progresso,
                        callback_recurso_concluido=recurso_status_callback,
                        callback_lista_recursos=lista_recursos_callback,
                        callback_cancelar=lambda: cancelar_flag,
                        caminho_planilha=caminho_usr,
                    )

                    def _ok_unico_v2():
                        sucesso_lst = resultado.get("recursos_sucesso", [])
                        falha_lst   = resultado.get("recursos_falha",   [])

                        txt_suc = "\n".join([f"  • {r}" for r in sucesso_lst]) if sucesso_lst else "  (Nenhum)"
                        txt_fal = "\n".join([f"  • {r}" for r in falha_lst])   if falha_lst   else "  (Nenhum)"

                        adicionar_log(
                            f"\n============================================================\n"
                            f"📊 RELATÓRIO DA EXECUÇÃO — {data_ini} até {data_fim or data_ini}\n"
                            f"============================================================\n\n"
                            f"Recursos Concluídos com Sucesso ({len(sucesso_lst)}):\n"
                            f"{txt_suc}\n\n"
                            f"Recursos Sem Seriais / Pulados ({len(falha_lst)}):\n"
                            f"{txt_fal}\n\n"
                            f"------------------------------------------------------------\n"
                            f"Total de Seriais Enviados: {resultado['total_seriais']}\n"
                            f"Total de Lotes Executados: {resultado['total_lotes']}\n\n"
                            f"Log salvo em:\n{resultado['log']}\n"
                            f"============================================================\n"
                        )

                    btn_iniciar.after(0, _ok_unico_v2)








            except Exception as exc:
                err_msg = str(exc)

                def _erro():
                    adicionar_log(f"\nERRO: {err_msg}")
                    messagebox.showerror("Erro na automacao", err_msg)

                btn_iniciar.after(0, _erro)
            finally:
                set_btn(True)

                def _reset():
                    lbl_progresso.config(text="Aguardando...")

                btn_iniciar.after(0, _reset)

        threading.Thread(target=_executar, daemon=True).start()

    btn_iniciar.config(command=on_iniciar)

    # Mensagem de boas-vindas no log
    txt_log.insert(tk.END, "Sistema pronto.\n")
    txt_log.insert(tk.END, "Preencha os campos e clique em 'INICIAR AUTOMACAO'.\n")

    return frame_principal